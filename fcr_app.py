# ---------- fcr_crct.py ----------
import cv2
import numpy as np
import pandas as pd
import os
import subprocess
import sys
from ultralytics import YOLO
from datetime import datetime

print("🦐 Shrimp Detection & FCR Dashboard (Console Mode)\n")

# --- AUTO-INSTALL openpyxl IF MISSING ---
try:
    import openpyxl
    from openpyxl import load_workbook
except ImportError:
    print("⚙️ 'openpyxl' not found. Installing automatically...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl import load_workbook
    print("✅ 'openpyxl' installed successfully!\n")

# --- PATH CONFIG ---
model_path = r"F:\shrimp_model\best.pt"       # YOLO model path
input_folder = r"F:\shrimp_images"            # Folder containing shrimp images
output_folder = r"F:\shrimp_result"           # Folder to save results
os.makedirs(output_folder, exist_ok=True)

# --- LOAD YOLO MODEL ---
if not os.path.exists(model_path):
    print(f"❌ YOLO model not found at {model_path}")
    exit()

model = YOLO(model_path)
print("✅ YOLO model loaded successfully!")

# --- CONSTANTS ---
a, b = 0.0046, 2.99           # Length-weight relationship constants
ruler_length_mm = 300          # Ruler reference length (mm)

# --- GET IMAGES ---
image_files = [f for f in os.listdir(input_folder) if f.lower().endswith(('.jpg', '.jpeg', '.png'))]

if not image_files:
    print("📁 No shrimp images found in folder:", input_folder)
    exit()

data_records = []

# --- PROCESS IMAGES ---
for image_name in image_files:
    print(f"\n🔍 Processing: {image_name}")

    img_path = os.path.join(input_folder, image_name)
    img = cv2.imread(img_path)
    if img is None:
        print(f"⚠️ Could not read image: {image_name}")
        continue

    img_rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)

    # --- YOLO DETECTION ---
    results = model(img_rgb)
    result = results[0]
    boxes = result.boxes
    names = result.names

    ruler_boxes, shrimp_boxes = [], []

    for box in boxes:
        cls = int(box.cls[0])
        label = names[cls].lower()
        if label == "ruler":
            ruler_boxes.append(box.xyxy[0])
        elif label == "shrimp":
            shrimp_boxes.append(box.xyxy[0])

    if not ruler_boxes:
        print(f"⚠️ No ruler detected in {image_name}, skipping...")
        continue

    # --- PIXELS TO MILLIMETER CONVERSION ---
    x1, y1, x2, y2 = ruler_boxes[0]
    ruler_px_len = float(((x2 - x1)**2 + (y2 - y1)**2)**0.5)
    px_per_mm = ruler_px_len / ruler_length_mm

    shrimp_lengths, shrimp_weights = [], []

    for shrimp_box in shrimp_boxes:
        sx1, sy1, sx2, sy2 = map(int, shrimp_box)
        shrimp_px_len = float(((sx2 - sx1)**2 + (sy2 - sy1)**2)**0.5)
        shrimp_mm_len = shrimp_px_len / px_per_mm
        shrimp_cm_len = shrimp_mm_len / 10.0
        shrimp_weight_g = a * (shrimp_cm_len ** b)

        shrimp_lengths.append(round(shrimp_mm_len, 2))
        shrimp_weights.append(round(shrimp_weight_g, 2))

        # Draw bounding box and text
        cv2.rectangle(img, (sx1, sy1), (sx2, sy2), (0, 255, 0), 2)
        cv2.putText(
            img,
            f"{shrimp_mm_len:.1f}mm | {shrimp_weight_g:.2f}g",
            (sx1, max(30, sy1 - 10)),
            cv2.FONT_HERSHEY_SIMPLEX,
            0.7,
            (255, 255, 255),
            2,
            cv2.LINE_AA,
        )

    # --- SUMMARY CALCULATION ---
    shrimp_count = len(shrimp_boxes)
    avg_length = np.mean(shrimp_lengths) if shrimp_lengths else 0
    avg_weight = np.mean(shrimp_weights) if shrimp_weights else 0
    total_biomass = avg_weight * shrimp_count

    # --- FEED INPUT & FCR ---
    feed_input = 100  # grams (example value)
    weight_gain = total_biomass
    fcr = round(feed_input / weight_gain, 2) if weight_gain > 0 else 0

    # --- FEED INPUT SUGGESTION PLACEHOLDER ---
    feed_suggestion = 0  # will be Excel formula

    # --- SAVE DATA ---
    data_records.append({
        "Image": image_name,
        "Individual_Lengths_mm": shrimp_lengths,
        "Individual_Weights_g": shrimp_weights,
        "Avg_Length_mm": round(avg_length, 2),
        "Avg_Weight_g": round(avg_weight, 2),
        "Feed_Input_g": feed_input,
        "Weight_Gain_g": round(weight_gain, 2),
        "FCR": fcr,  # placeholder, will be formula
        "Feed_Input_Suggestion_g": feed_suggestion  # placeholder
    })

    # --- SAVE ANNOTATED IMAGE ---
    save_path = os.path.join(output_folder, image_name)
    cv2.imwrite(save_path, img)
    print(f"✅ Saved annotated image to: {save_path}")

# --- EXPORT RESULTS TO EXCEL (with live formulas) ---
df = pd.DataFrame(data_records)
excel_filename = f"shrimp_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
excel_path = os.path.join(output_folder, excel_filename)
df.to_excel(excel_path, index=False)

# Reopen Excel to insert formulas
wb = load_workbook(excel_path)
ws = wb.active

# Get column positions
headers = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}
feed_col = headers.get("Feed_Input_g")
weight_gain_col = headers.get("Weight_Gain_g")
fcr_col = headers.get("FCR")
suggest_col = headers.get("Feed_Input_Suggestion_g")

if feed_col and weight_gain_col and fcr_col and suggest_col:
    for row in range(2, ws.max_row + 1):
        # Excel formula for FCR
        ws.cell(row=row, column=fcr_col).value = f"=ROUND({chr(64+feed_col)}{row}/{chr(64+weight_gain_col)}{row},2)"
        # Excel formula for feed suggestion (in grams)
        ws.cell(row=row, column=suggest_col).value = (
            f"=IF({chr(64+fcr_col)}{row}>2,"
            f"{chr(64+feed_col)}{row}*0.9,"
            f"IF({chr(64+fcr_col)}{row}<1.2,"
            f"{chr(64+feed_col)}{row}*1.1,"
            f"{chr(64+feed_col)}{row}))"
        )

wb.save(excel_path)

print("\n📊 Detection Summary:")
print(df)
print(f"\n💾 Results saved to Excel (with live FCR & Feed Suggestion formulas): {excel_path}")

# --- OPEN EXCEL FILE AUTOMATICALLY ---
try:
    os.startfile(excel_path)
    print("📂 Excel file opened automatically.")
except Exception:
    print("⚠️ Could not open Excel automatically. Please open it manually.")
